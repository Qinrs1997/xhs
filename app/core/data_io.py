"""数据导入导出模块

提供通用的 Excel/CSV 数据导入导出功能。

使用方法：
    from app.core.data_io import DataExporter, DataImporter

    # 导出为 CSV
    exporter = DataExporter()
    response = await exporter.export_csv(data, columns, filename="users.csv")

    # 导出为 Excel
    response = await exporter.export_excel(data, columns, filename="users.xlsx")

    # 导入 CSV
    importer = DataImporter()
    records = await importer.import_csv(file, required_columns=["username", "email"])

    # 导入 Excel
    records = await importer.import_excel(file, required_columns=["username", "email"])
"""
import csv
import io
from typing import Any, Optional
from datetime import datetime

from fastapi import UploadFile
from fastapi.responses import StreamingResponse

from app.core.logger import logger


# Excel 支持（可选依赖）
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class DataExporter:
    """数据导出器

    支持将数据导出为 CSV 或 Excel 格式的 StreamingResponse。
    """

    @staticmethod
    def _serialize_value(value: Any) -> str:
        """将值序列化为字符串"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, bool):
            return "是" if value else "否"
        return str(value)

    async def export_csv(
        self,
        data: list[dict[str, Any]],
        columns: dict[str, str],
        filename: str = "export.csv",
    ) -> StreamingResponse:
        """导出为 CSV 文件

        Args:
            data: 数据列表（字典格式）
            columns: 列映射 {字段名: 显示名}，例如 {"username": "用户名", "email": "邮箱"}
            filename: 下载文件名

        Returns:
            StreamingResponse（可直接作为 API 返回值）
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # 写入表头
        headers = list(columns.values())
        writer.writerow(headers)

        # 写入数据
        field_names = list(columns.keys())
        for row in data:
            writer.writerow([
                self._serialize_value(row.get(field, ""))
                for field in field_names
            ])

        # 转为 bytes（带 BOM 头，确保 Excel 正确识别 UTF-8）
        content = "\ufeff" + output.getvalue()
        output.close()

        return StreamingResponse(
            iter([content.encode("utf-8")]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )

    async def export_excel(
        self,
        data: list[dict[str, Any]],
        columns: dict[str, str],
        filename: str = "export.xlsx",
        sheet_name: str = "Sheet1",
    ) -> StreamingResponse:
        """导出为 Excel 文件

        Args:
            data: 数据列表（字典格式）
            columns: 列映射 {字段名: 显示名}
            filename: 下载文件名
            sheet_name: 工作表名称

        Returns:
            StreamingResponse

        Raises:
            ImportError: 未安装 openpyxl
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                "Excel 导出需要安装 openpyxl: pip install openpyxl"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写入表头 + 样式
        headers = list(columns.values())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

        # 写入数据
        field_names = list(columns.keys())
        for row in data:
            ws.append([
                self._serialize_value(row.get(field, ""))
                for field in field_names
            ])

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:  # noqa: S110 -- 仅用于列宽估算，单格失败无需打日志
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

        # 写入内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )


class DataImporter:
    """数据导入器

    支持从 CSV 或 Excel 文件导入数据，提供格式校验和错误报告。
    """

    async def import_csv(
        self,
        file: UploadFile,
        required_columns: Optional[list[str]] = None,
        max_rows: int = 10000,
    ) -> dict[str, Any]:
        """从 CSV 文件导入数据

        Args:
            file: 上传的 CSV 文件
            required_columns: 必须存在的列名列表
            max_rows: 最大行数限制

        Returns:
            {
                "columns": [...],     # 列名列表
                "records": [...],     # 数据记录（字典列表）
                "total": int,         # 总行数
                "errors": [...]       # 错误信息列表
            }
        """
        content = await file.read()
        errors = []

        # 尝试不同编码
        for encoding in ["utf-8-sig", "utf-8", "gbk", "gb2312"]:
            try:
                text = content.decode(encoding)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            return {
                "columns": [],
                "records": [],
                "total": 0,
                "errors": ["无法识别文件编码，请使用 UTF-8 编码"],
            }

        reader = csv.DictReader(io.StringIO(text))
        columns = reader.fieldnames or []

        # 校验必须列
        if required_columns:
            missing = [c for c in required_columns if c not in columns]
            if missing:
                return {
                    "columns": columns,
                    "records": [],
                    "total": 0,
                    "errors": [f"缺少必须列: {', '.join(missing)}"],
                }

        records = []
        for i, row in enumerate(reader, start=2):
            if i > max_rows + 1:
                errors.append(f"数据超过 {max_rows} 行限制，已截断")
                break
            # 清理空白
            cleaned = {k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
            records.append(cleaned)

        logger.info("CSV 导入完成: {} 行, {} 个错误", len(records), len(errors))

        return {
            "columns": columns,
            "records": records,
            "total": len(records),
            "errors": errors,
        }

    async def import_excel(
        self,
        file: UploadFile,
        required_columns: Optional[list[str]] = None,
        max_rows: int = 10000,
        sheet_name: Optional[str] = None,
    ) -> dict[str, Any]:
        """从 Excel 文件导入数据

        Args:
            file: 上传的 Excel 文件
            required_columns: 必须存在的列名列表
            max_rows: 最大行数限制
            sheet_name: 指定工作表名（默认取第一个）

        Returns:
            与 import_csv 相同的格式
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                "Excel 导入需要安装 openpyxl: pip install openpyxl"
            )

        content = await file.read()
        errors = []

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
        except Exception as e:
            return {
                "columns": [],
                "records": [],
                "total": 0,
                "errors": [f"无法读取 Excel 文件: {e}"],
            }

        # 选择工作表
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return {
                    "columns": [],
                    "records": [],
                    "total": 0,
                    "errors": [f"工作表 '{sheet_name}' 不存在"],
                }
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {
                "columns": [],
                "records": [],
                "total": 0,
                "errors": ["文件为空"],
            }

        # 第一行为表头
        columns = [str(c).strip() if c else f"列{i}" for i, c in enumerate(rows[0], 1)]

        # 校验必须列
        if required_columns:
            missing = [c for c in required_columns if c not in columns]
            if missing:
                return {
                    "columns": columns,
                    "records": [],
                    "total": 0,
                    "errors": [f"缺少必须列: {', '.join(missing)}"],
                }

        records = []
        for i, row in enumerate(rows[1:], start=2):
            if i > max_rows + 1:
                errors.append(f"数据超过 {max_rows} 行限制，已截断")
                break

            record = {}
            for j, value in enumerate(row):
                if j < len(columns):
                    if isinstance(value, str):
                        value = value.strip()
                    record[columns[j]] = value
            records.append(record)

        wb.close()

        logger.info("Excel 导入完成: {} 行, {} 个错误", len(records), len(errors))

        return {
            "columns": columns,
            "records": records,
            "total": len(records),
            "errors": errors,
        }


# 全局实例
data_exporter = DataExporter()
data_importer = DataImporter()
